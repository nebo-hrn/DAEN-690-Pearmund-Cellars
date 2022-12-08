#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Oct 14 14:22:26 2022

@author: nebojsahrnjez
"""
#%% Import Statments
# Standard imports
import pandas as pd
import numpy as np
import datetime
from tqdm import tqdm
from sklearn.preprocessing import MinMaxScaler
import scipy.stats as st
from dateutil.relativedelta import relativedelta


# Excel File Imports
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# ARIMA models
import pmdarima as pm

#%% Prevent Arima Output

import sys

class NullWriter(object):
    def write(self, arg):
        pass
    
nullwrite = NullWriter()
oldstdout = sys.stdout

#%%Reading in Excel and Getting Dataframes

path = "C:/Users/Nebojsa/Desktop/DAEN690/Capstone_Excel_Format.xlsm"

df = pd.read_excel(path, sheet_name=None, engine="openpyxl", parse_dates=True)

PC = df["PC Inventory"]
EF = df["EF Inventory"]
VH = df["VH Inventory"]

winery = df['Home'].iloc[23,5]


PC["Winery"] = "PC"
PC["Cases Moved"] = pd.to_numeric(PC["Cases Moved"])
for i in range(len(PC)):
    PC.loc[i,["Sale Date"]] = datetime.datetime(int(PC.loc[i,['Fiscal Year']]), int(PC.loc[i,['Month']]), 1)
EF["Winery"] = "EF"
EF["Cases Moved"] = pd.to_numeric(EF["Cases Moved"])
for i in range(len(EF)):
    EF.loc[i,["Sale Date"]] = datetime.datetime(int(EF.loc[i,['Fiscal Year']]), int(EF.loc[i,['Month']]), 1)
VH["Winery"] = "VH"
VH["Cases Moved"] = pd.to_numeric(VH["Cases Moved"])
for i in range(len(VH)):
    VH.loc[i,["Sale Date"]] = datetime.datetime(int(VH.loc[i,['Fiscal Year']]), int(VH.loc[i,['Month']]), 1)

#frames = [PC, EF, VH]
#m_df = pd.concat(frames)

#%%Grouping by sale date and wine type, summing cases moved

PC = PC.groupby(["Sale Date","Winery", "Wine"])["Cases Moved"].sum().reset_index()
EF = EF.groupby(["Sale Date","Winery", "Wine"])["Cases Moved"].sum().reset_index()
VH = VH.groupby(["Sale Date","Winery", "Wine"])["Cases Moved"].sum().reset_index()

#%% Modeling
if winery == "Pearmund Cellars":
    wineries = ["PC"]
    df = PC
elif winery == "Effingham Manor":
    wineries = ["EF"]
    df = EF
elif winery == "Vint Hill":
    wineries = ["VH"]
    df = VH
elif winery == "PC + EF":
    wineries = ["PC","EF"]
    frames = [PC, EF]
    df = pd.concat(frames)
    

data_wine_list = [i for i in df["Wine"].unique()]

good_wine_list_txt = open("C:/Users/Nebojsa/Desktop/DAEN690/Tracked Wine List.txt", "r")
wine_list_data = good_wine_list_txt.read()
good_wine_list = wine_list_data.split("\\")
good_wine_list_txt.close
good_wine_list = [i.rstrip() for i in good_wine_list]
good_wine_list = [i.lstrip() for i in good_wine_list]

wine_list = []
for i in data_wine_list:
    if i in good_wine_list:
        wine_list.append(i)
    else:
        continue
# Predict only for the relevant wines.


predict = pd.DataFrame()
# empty datafraime to hold all values
month = datetime.datetime(2022,10,1)
month2 = month - relativedelta(months=36)

print("Predicting " + winery + "\n" )
for j in tqdm(wineries, desc = "Wineries"):
    for i in tqdm(wine_list, desc = "Wines", leave = None):
        
        sys.stdout = nullwrite
        
        data = df[(df.Winery == j) & (df.Wine == i) & (df["Sale Date"] < month)]
        test = data[data["Sale Date"] > month2]
        if len(test) < 30:
            continue
        # Skip wines that have less than 3 of recent years of historical data
        else:
            model = pm.auto_arima(
                data["Cases Moved"].values,
                start_p=0,
                start_q=0,
                test="adf",
                max_p=20,
                max_q=20,
                m=12,
                d=None,
                seasonal=True,
                start_P=0,
                D=1,
                trace=True,
                error_action="ignore",
                suppress_warnings=True,
                stepwise=True,
                with_intercept=True,
                **{"disp":0}
            )
    
            # Forecast
            n_periods = 12
            fc, confint = model.predict(n_periods=n_periods, return_conf_int=True)
            index_of_fc = np.arange(
                len(data["Cases Moved"].values), len(data["Cases Moved"].values) + n_periods
            )
            # Forecast 12 months into the future, anything else is basically useless
    
            # make series for plotting purpose
            fc_series = pd.Series(fc)
            lower_series = pd.Series(confint[:, 0], index=index_of_fc)
            upper_series = pd.Series(confint[:, 1], index=index_of_fc)
    
            val_range = st.t.interval(
                0.7,
                len(data["Cases Moved"].values) - 1,
                loc=np.mean(data["Cases Moved"].values),
                scale=st.sem(data["Cases Moved"].values),
            )
            # Find the 80% confidence interval of the historical data, this is later used for scaling
    
            if val_range[0] < 0:
                val_low = 0
            else:
                val_low = val_range[0]
            val_high = val_range[1]
            # Set high and low range as 80% confidence interval of historical data, if low is 0 for
            # some reason, change to 0
    
            if any(fc_series < 0) == True:
                scaler = MinMaxScaler(feature_range=(val_low, val_high))
                scaler.fit(np.array(fc_series.values).reshape(-1, 1))
                fc_series = scaler.transform(np.array(fc_series.values).reshape(-1, 1))
            else:
                fc_series = np.array(fc_series.values).reshape(-1, 1)
            # If any prediction values are negative, scale them so they are positive
            
            if len(wineries) > 1:
                wine_name = j + " - " + i
            else:
                wine_name = i
    
            demand = pd.DataFrame(
                {
                    "Wine": wine_name,
                    "Month 1": fc_series[0][0],
                    "Month 2": fc_series[1][0],
                    "Month 3": fc_series[2][0],
                    "Month 4": fc_series[3][0],
                    "Month 5": fc_series[4][0],
                    "Month 6": fc_series[5][0],
                    "Month 7": fc_series[6][0],
                    "Month 8": fc_series[7][0],
                    "Month 9": fc_series[8][0],
                    "Month 10": fc_series[9][0],
                    "Month 11": fc_series[10][0],
                    "Month 12": fc_series[11][0],
                },
                index=[0],
            )
            # Fill a dataframe with the predicted values so it can be added to a prediction dataframe
            # and later written to an excel sheet
    
            predict = pd.concat([predict, demand])
            # Add that wine to the larger prediction dataframe
            
            
sys.stdout = oldstdout
predict["Total"] = predict.sum(axis=1, numeric_only=True)
# Summing all the months to produce a yearly total

print("Prediction Complete")

#%%
wb = load_workbook("C:/Users/Nebojsa/Desktop/DAEN690/Demand.xlsx")
sheet = wb['Demand']


for row in sheet["A1:Z56"]:
    for cell in row:
        cell.value = ""


rows = dataframe_to_rows(predict, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx, column=c_idx, value=value)
# Write the prediction to the sheet


for i in range(1, r_idx + 1):
    for j in range(1, c_idx + 1):
        if i == 1:
            sheet.cell(row=i, column=j).font = Font(bold=True)
            sheet.cell(row=i, column=j).border = Border(
                top=Side(border_style="thick"), bottom=Side(border_style="thick")
            )

        if j == 1:
            sheet.cell(row=i, column=j).font = Font(bold=True)
            sheet.cell(row=i, column=j).border = Border(
                left=Side(border_style="thick"), right=Side(border_style="thick")
            )

        if j == c_idx:
            sheet.cell(row=i, column=j).border = Border(
                right=Side(border_style="thick"), left=Side(border_style="thick")
            )

        if i == r_idx:
            sheet.cell(row=i, column=j).border = Border(bottom=Side(border_style="thick"))

        if (i == 1) & (j == c_idx):
            sheet.cell(row=i, column=j).border = Border(
                right=Side(border_style="thick"),
                bottom=Side(border_style="thick"),
                top=Side(border_style="thick"),
            )

        if (i == r_idx) & (j == c_idx):
            sheet.cell(row=i, column=j).border = Border(
                right=Side(border_style="thick"),
                bottom=Side(border_style="thick"),
                left=Side(border_style="thick"),
            )

        if (i == r_idx) & (j == 1):
            sheet.cell(row=i, column=j).border = Border(
                right=Side(border_style="thick"),
                bottom=Side(border_style="thick"),
                left=Side(border_style="thick"),
            )

        if (i == 1) & (j == 1):
            sheet.cell(row=i, column=j).border = Border(
                right=Side(border_style="thick"),
                bottom=Side(border_style="thick"),
                top=Side(border_style="thick"),
                left=Side(border_style="thick"),
            )


wb.save("C:/Users/Nebojsa/Desktop/DAEN690/Demand.xlsx")
wb.close()


#%%Test
